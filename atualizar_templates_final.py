#!/usr/bin/env python
"""
Script para atualizar os templates e garantir que estejam corretos
"""
import os
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def criar_template_empresas():
    """Cria template de empresas com código"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"
    
    # Cabeçalhos
    headers = [
        'Código',
        'Razão Social',
        'CNPJ',
        'Nome Fantasia',
        'E-mail',
        'Telefone',
        'Endereço',
        'Responsável'
    ]
    
    # Adicionar cabeçalhos
    ws.append(headers)
    
    # Adicionar linha de exemplo
    ws.append([
        'EMP001',
        'Empresa Exemplo Ltda',
        '12345678000190',
        'Exemplo',
        'contato@exemplo.com',
        '(11) 99999-9999',
        'Rua Exemplo, 123 - São Paulo - SP',
        'João Silva'
    ])
    
    # Formatação
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Aplicar formatação ao cabeçalho
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Ajustar largura das colunas
    column_widths = [15, 40, 18, 30, 30, 18, 40, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Congelar primeira linha
    ws.freeze_panes = "A2"
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:{chr(65 + len(headers) - 1)}1"
    
    # Salvar
    output_path = os.path.join('backend', 'media', 'templates', 'template_empresas.xlsx')
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    
    print(f"✅ Template de empresas criado: {output_path}")
    return output_path

def criar_template_obrigacoes():
    """Cria template de obrigações com CNPJ"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Obrigações"
    
    # Cabeçalhos
    headers = [
        'CNPJ da Empresa',
        'Estado',
        'Tipo de Obrigação',
        'Nome da Obrigação',
        'Competência',
        'Data de Vencimento',
        'Prazo de Entrega',
        'Usuário Responsável',
        'Data Inicial de Validade',
        'Data Final de Validade',
        'Notas'
    ]
    
    # Adicionar cabeçalhos
    ws.append(headers)
    
    # Adicionar linha de exemplo
    ws.append([
        '12345678000190',
        'SP',
        'SPED Fiscal',
        'SPED Fiscal',
        '01/2024',
        '2024-01-31',
        '2024-01-25',
        'admin',
        '2024-01-01',
        '2024-12-31',
        'Obrigação mensal'
    ])
    
    # Formatação
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Aplicar formatação ao cabeçalho
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Ajustar largura das colunas
    column_widths = [18, 8, 30, 30, 12, 15, 15, 20, 15, 15, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Congelar primeira linha
    ws.freeze_panes = "A2"
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:{chr(65 + len(headers) - 1)}1"
    
    # Salvar
    output_path = os.path.join('backend', 'media', 'templates', 'template_obrigacoes.xlsx')
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    
    print(f"✅ Template de obrigações criado: {output_path}")
    return output_path

def main():
    print("=" * 70)
    print("ATUALIZANDO TEMPLATES - VERSÃO FINAL")
    print("=" * 70)
    print()
    
    # Criar templates
    template_empresas = criar_template_empresas()
    template_obrigacoes = criar_template_obrigacoes()
    
    print()
    print("=" * 70)
    print("✅ TEMPLATES ATUALIZADOS COM SUCESSO!")
    print("=" * 70)
    print()
    print("📄 Templates criados:")
    print(f"   1. {template_empresas}")
    print(f"   2. {template_obrigacoes}")
    print()
    print("📝 IMPORTANTE:")
    print("   1. Reinicie o servidor Django (Ctrl+C e depois python manage.py runserver)")
    print("   2. Limpe o cache do navegador (Ctrl+Shift+Del)")
    print("   3. Baixe o template novamente")
    print()
    print("🔍 Verifique se o template de empresas tem:")
    print("   - Primeira coluna: 'Código'")
    print("   - Segunda coluna: 'Razão Social'")
    print("   - Terceira coluna: 'CNPJ'")
    print()
    print("🔍 Verifique se o template de obrigações tem:")
    print("   - Primeira coluna: 'CNPJ da Empresa'")
    print("   - Segunda coluna: 'Estado'")
    print("   - Terceira coluna: 'Tipo de Obrigação'")
    print()

if __name__ == '__main__':
    main()

