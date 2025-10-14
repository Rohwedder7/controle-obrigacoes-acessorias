from django.core.management.base import BaseCommand
from openpyxl import Workbook
import os
from django.conf import settings

class Command(BaseCommand):
    help = 'Cria templates de planilha para upload em massa'

    def handle(self, *args, **options):
        # Criar diretório de templates se não existir
        templates_dir = os.path.join(settings.MEDIA_ROOT, 'templates')
        os.makedirs(templates_dir, exist_ok=True)
        
        # Template para empresas
        self.create_companies_template(templates_dir)
        
        # Template para obrigações
        self.create_obligations_template(templates_dir)
        
        self.stdout.write(
            self.style.SUCCESS('Templates criados com sucesso!')
        )

    def create_companies_template(self, templates_dir):
        """Cria template para upload de empresas"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Empresas"
        
        # Cabeçalhos
        headers = [
            'Nome da Empresa',
            'CNPJ',
            'Nome Fantasia',
            'E-mail',
            'Telefone',
            'Endereço',
            'Responsável'
        ]
        
        # Aplicar estilo aos cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = cell.font.copy(bold=True)
            cell.fill = cell.fill.copy(fill_type='solid', start_color='CCCCCC')
        
        # Exemplos de dados
        examples = [
            ['Empresa Exemplo Ltda', '12.345.678/0001-90', 'Exemplo Corp', 'contato@exemplo.com', '(11) 99999-9999', 'Rua Exemplo, 123', 'João Silva'],
            ['Outra Empresa S.A.', '98.765.432/0001-10', 'Outra Corp', 'contato@outra.com', '(11) 88888-8888', 'Av. Outra, 456', 'Maria Santos'],
        ]
        
        for row, example in enumerate(examples, 2):
            for col, value in enumerate(example, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Ajustar largura das colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
        
        # Salvar arquivo
        file_path = os.path.join(templates_dir, 'template_empresas.xlsx')
        wb.save(file_path)
        self.stdout.write(f'Template de empresas criado: {file_path}')

    def create_obligations_template(self, templates_dir):
        """Cria template para upload de obrigações"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Obrigações"
        
        # Cabeçalhos
        headers = [
            'Nome da Empresa',
            'Estado (Código)',
            'Tipo de Obrigação',
            'Nome da Obrigação Acessória',
            'Competência (MM/AAAA)',
            'Data de Vencimento (YYYY-MM-DD)',
            'Prazo de Entrega (YYYY-MM-DD)',
            'Usuário Responsável (Username)',
            'Data Inicial de Validade (YYYY-MM-DD)',
            'Data Final de Validade (YYYY-MM-DD)',
            'Observações'
        ]
        
        # Aplicar estilo aos cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = cell.font.copy(bold=True)
            cell.fill = cell.fill.copy(fill_type='solid', start_color='CCCCCC')
        
        # Exemplos de dados
        examples = [
            [
                'Empresa Exemplo Ltda',
                'SP',
                'Federal',
                'SPED - Escrituração Fiscal Digital',
                '01/2024',
                '2024-01-31',
                '2024-02-15',
                'admin',
                '2024-01-01',
                '2024-12-31',
                'Obrigação mensal'
            ],
            [
                'Outra Empresa S.A.',
                'RJ',
                'Estadual',
                'ICMS',
                '01/2024',
                '2024-02-28',
                '2024-03-15',
                'admin',
                '2024-01-01',
                '2024-12-31',
                'Obrigação mensal'
            ],
            [
                'Terceira Empresa Ltda',
                'MG',
                'Municipal',
                'ISS',
                '01/2024',
                '2024-02-25',
                '2024-03-10',
                'admin',
                '2024-01-01',
                '2024-12-31',
                'Obrigação mensal'
            ],
            [
                'Quarta Empresa S.A.',
                'PR',
                'Regimes Especiais',
                'DAS - Simples Nacional',
                '01/2024',
                '2024-02-20',
                '2024-03-05',
                'admin',
                '2024-01-01',
                '2024-12-31',
                'Obrigação mensal'
            ],
        ]
        
        for row, example in enumerate(examples, 2):
            for col, value in enumerate(example, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Ajustar largura das colunas
        column_widths = [25, 15, 20, 30, 15, 20, 20, 20, 25, 25, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Adicionar aba de instruções
        ws_instructions = wb.create_sheet("Instruções")
        instructions = [
            ["INSTRUÇÕES PARA PREENCHIMENTO"],
            [""],
            ["1. Nome da Empresa: Nome completo da empresa (deve existir no sistema)"],
            ["2. Estado (Código): Código do estado (ex: SP, RJ, MG, PR)"],
            ["3. Tipo de Obrigação: Federal, Estadual, Municipal ou Regimes Especiais"],
            ["4. Nome da Obrigação Acessória: Nome específico (ex: SPED, ICMS, ISS, DAS)"],
            ["5. Competência: Mês/Ano no formato MM/AAAA (ex: 01/2024)"],
            ["6. Data de Vencimento: Data no formato YYYY-MM-DD (ex: 2024-01-31)"],
            ["7. Prazo de Entrega: Data no formato YYYY-MM-DD (opcional)"],
            ["8. Usuário Responsável: Username do usuário (deve existir no sistema)"],
            ["9. Data Inicial de Validade: Data no formato YYYY-MM-DD (opcional)"],
            ["10. Data Final de Validade: Data no formato YYYY-MM-DD (opcional)"],
            ["11. Observações: Campo livre para observações"],
            [""],
            ["IMPORTANTE: Não altere os cabeçalhos das colunas!"],
        ]
        
        for row, instruction in enumerate(instructions, 1):
            ws_instructions.cell(row=row, column=1, value=instruction[0])
            if row == 1:  # Título
                ws_instructions.cell(row=row, column=1).font = ws_instructions.cell(row=row, column=1).font.copy(bold=True, size=14)
        
        # Salvar arquivo
        file_path = os.path.join(templates_dir, 'template_obrigacoes.xlsx')
        wb.save(file_path)
        self.stdout.write(f'Template de obrigações criado: {file_path}')
