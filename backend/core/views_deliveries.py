import re
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from rest_framework.decorators import api_view, permission_classes
from rest_framework import permissions
from rest_framework.response import Response
from rest_framework import status
from .models import Company, Obligation, Submission, SubmissionAttachment, AuditLog
from .serializers import ObligationSerializer


def audit(user, action, obj, changes=None):
    """Registrar evento no audit log"""
    AuditLog.objects.create(
        user=user,
        action=action,
        model=obj.__class__.__name__,
        object_id=obj.id if hasattr(obj, 'id') else None,
        changes=changes or {}
    )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_company_obligations(request, company_id):
    """
    Listar obrigações de uma empresa para select dependente
    GET /api/companies/{id}/obligations/
    """
    try:
        company = Company.objects.get(id=company_id)
        obligations = Obligation.objects.filter(company=company).select_related(
            'state', 'obligation_type'
        ).order_by('-competence', 'obligation_type__name')
        
        data = []
        for obligation in obligations:
            label = f"{obligation.obligation_name or obligation.obligation_type.name} • {obligation.state.code} • {obligation.competence}"
            data.append({
                'id': obligation.id,
                'label': label,
                'competence': obligation.competence,
                'due_date': obligation.due_date.isoformat() if obligation.due_date else None,
                'has_submission': obligation.submissions.exists()
            })
        
        return Response(data)
    except Company.DoesNotExist:
        return Response({'error': 'Empresa não encontrada'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def download_delivery_template(request):
    """
    Baixar template Excel para entregas em massa
    GET /api/deliveries/template.xlsx
    """
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Entregas"
    
    # Cabeçalhos
    headers = [
        'cnpj', 'company_name', 'state', 'obligation_name', 
        'competence', 'delivery_date', 'type', 'comments'
    ]
    
    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escrever cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Exemplos de dados
    examples = [
        ['12345678000190', 'Empresa Exemplo Ltda', 'RS', 'EFD Contribuições', '08/2025', '2025-08-15', 'original', 'Entrega original'],
        ['12345678000190', '', 'RS', 'EFD Contribuições', '08/2025', '2025-08-20', 'retificadora', 'Retificação'],
        ['98765432000123', 'Outra Empresa S/A', 'SP', 'SPED Fiscal', '09/2025', '2025-09-10', 'original', '']
    ]
    
    for row, example in enumerate(examples, 2):
        for col, value in enumerate(example, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Comentários nas células (removido por incompatibilidade com openpyxl)
    # Os comentários podem ser adicionados manualmente se necessário
    
    # Ajustar largura das colunas
    column_widths = [15, 25, 8, 20, 12, 12, 12, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="template_entregas.xlsx"'
    
    wb.save(response)
    return response


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def bulk_deliveries(request):
    """
    Processar entregas em massa via planilha Excel
    POST /api/deliveries/bulk/
    """
    if 'file' not in request.FILES:
        return Response({'error': 'Arquivo não fornecido'}, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    if not file.name.endswith('.xlsx'):
        return Response({'error': 'Arquivo deve ser .xlsx'}, status=status.HTTP_400_BAD_REQUEST)
    
    batch_id = uuid.uuid4()
    created = 0
    updated = 0
    skipped = []
    
    try:
        # Ler planilha
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # Pular cabeçalho
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        with transaction.atomic():
            for row_num, row in enumerate(rows, 2):
                if not any(row):  # Linha vazia
                    continue
                
                try:
                    cnpj, company_name, state, obligation_name, competence, delivery_date, submission_type, comments = row[:8]
                    
                    # Validar dados obrigatórios
                    if not all([cnpj, state, obligation_name, competence, delivery_date]):
                        skipped.append({
                            'row': row_num,
                            'reason': 'Dados obrigatórios faltando',
                            'data': str(row)
                        })
                        continue
                    
                    # Limpar CNPJ (apenas números)
                    cnpj_clean = re.sub(r'\D', '', str(cnpj))
                    if len(cnpj_clean) != 14:
                        skipped.append({
                            'row': row_num,
                            'reason': 'CNPJ inválido',
                            'data': str(cnpj)
                        })
                        continue
                    
                    # Buscar empresa
                    try:
                        company = Company.objects.get(cnpj__contains=cnpj_clean)
                    except Company.DoesNotExist:
                        skipped.append({
                            'row': row_num,
                            'reason': 'Empresa não encontrada',
                            'data': f'CNPJ: {cnpj_clean}'
                        })
                        continue
                    
                    # Buscar obrigação
                    try:
                        obligation = Obligation.objects.get(
                            company=company,
                            state__code=state.upper(),
                            obligation_name__icontains=obligation_name
                        )
                    except Obligation.DoesNotExist:
                        skipped.append({
                            'row': row_num,
                            'reason': 'Obrigação não encontrada',
                            'data': f'{obligation_name} - {state}'
                        })
                        continue
                    
                    # Validar data
                    try:
                        from datetime import datetime
                        delivery_date_obj = datetime.strptime(str(delivery_date), '%Y-%m-%d').date()
                    except ValueError:
                        skipped.append({
                            'row': row_num,
                            'reason': 'Data inválida',
                            'data': str(delivery_date)
                        })
                        continue
                    
                    # Validar tipo
                    submission_type = str(submission_type).lower()
                    if submission_type not in ['original', 'retificadora']:
                        submission_type = 'original'
                    
                    # Buscar submission existente
                    existing_submission = Submission.objects.filter(
                        obligation=obligation
                    ).first()
                    
                    if existing_submission:
                        if submission_type == 'retificadora':
                            # Criar nova submission retificadora
                            submission = Submission.objects.create(
                                obligation=obligation,
                                delivered_by=request.user,
                                delivery_date=delivery_date_obj,
                                comments=str(comments) if comments else '',
                                submission_type='retificadora',
                                batch_id=batch_id
                            )
                            created += 1
                        else:
                            # Atualizar submission existente - IMPORTANTE: usar a data informada pelo usuário
                            existing_submission.delivery_date = delivery_date_obj
                            existing_submission.comments = str(comments) if comments else ''
                            existing_submission.batch_id = batch_id
                            existing_submission.save()
                            updated += 1
                    else:
                        # Criar nova submission - IMPORTANTE: usar a data informada pelo usuário
                        submission = Submission.objects.create(
                            obligation=obligation,
                            delivered_by=request.user,
                            delivery_date=delivery_date_obj,
                            comments=str(comments) if comments else '',
                            submission_type=submission_type,
                            batch_id=batch_id
                        )
                        created += 1
                    
                    # Registrar no audit log
                    audit(request.user, 'delivery_bulk', submission, {
                        'batch_id': str(batch_id),
                        'row': row_num
                    })
                    
                except Exception as e:
                    skipped.append({
                        'row': row_num,
                        'reason': f'Erro: {str(e)}',
                        'data': str(row)
                    })
                    continue
    
    except Exception as e:
        return Response({'error': f'Erro ao processar arquivo: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    return Response({
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'batch_id': str(batch_id),
        'message': f'Processamento concluído: {created} criadas, {updated} atualizadas, {len(skipped)} puladas'
    })


def parse_filename(filename):
    """
    Parse filename para extrair CNPJ, período e nome da obrigação
    Formato: CNPJsemPontos_PeriodoSemSeparadores_NomeObrigacao.ext
    Exemplo: 12345678000190_082026_EFDContribuicoes.zip
    """
    # Regex para o padrão
    pattern = r'^(\d{14})_(\d{2})(\d{4})_([A-Za-z0-9_]+)\.(zip|pdf|rar)$'
    match = re.match(pattern, filename)
    
    if not match:
        return None
    
    cnpj, month, year, obligation_key, extension = match.groups()
    period = f"{month}{year}"  # MMAAAA
    
    return {
        'cnpj': cnpj,
        'period': period,
        'obligation_key': obligation_key,
        'extension': extension
    }


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def bulk_attachments(request):
    """
    Processar anexos em massa
    POST /api/deliveries/bulk-attachments/
    """
    if 'files' not in request.FILES:
        return Response({'error': 'Arquivos não fornecidos'}, status=status.HTTP_400_BAD_REQUEST)
    
    files = request.FILES.getlist('files')
    attachments_linked = 0
    skipped = []
    
    with transaction.atomic():
        for file in files:
            try:
                # Parse do nome do arquivo
                parsed = parse_filename(file.name)
                if not parsed:
                    skipped.append({
                        'filename': file.name,
                        'reason': 'Nome do arquivo não segue o padrão esperado'
                    })
                    continue
                
                cnpj = parsed['cnpj']
                period = parsed['period']
                obligation_key = parsed['obligation_key']
                
                # Buscar empresa
                try:
                    company = Company.objects.get(cnpj__contains=cnpj)
                except Company.DoesNotExist:
                    skipped.append({
                        'filename': file.name,
                        'reason': 'Empresa não encontrada',
                        'data': f'CNPJ: {cnpj}'
                    })
                    continue
                
                # Converter período para competência
                month = period[:2]
                year = period[2:]
                competence = f"{month}/{year}"
                
                # Buscar obrigação
                try:
                    obligation = Obligation.objects.get(
                        company=company,
                        competence=competence,
                        obligation_name__icontains=obligation_key
                    )
                except Obligation.DoesNotExist:
                    skipped.append({
                        'filename': file.name,
                        'reason': 'Obrigação não encontrada',
                        'data': f'{obligation_key} - {competence}'
                    })
                    continue
                
                # Buscar ou criar submission - usar data atual se não especificada
                submission, created = Submission.objects.get_or_create(
                    obligation=obligation,
                    defaults={
                        'delivered_by': request.user,
                        'delivery_date': timezone.now().date(),
                        'submission_type': 'original'
                    }
                )
                
                # Criar attachment
                attachment = SubmissionAttachment.objects.create(
                    submission=submission,
                    file=file,
                    original_filename=file.name,
                    parsed_cnpj=cnpj,
                    parsed_period=period,
                    parsed_obligation_key=obligation_key
                )
                
                attachments_linked += 1
                
                # Registrar no audit log
                audit(request.user, 'delivery_attachments_linked', attachment, {
                    'filename': file.name,
                    'submission_id': submission.id
                })
                
            except Exception as e:
                skipped.append({
                    'filename': file.name,
                    'reason': f'Erro: {str(e)}'
                })
                continue
    
    return Response({
        'attachments_linked': attachments_linked,
        'skipped': skipped,
        'message': f'Processamento concluído: {attachments_linked} anexos vinculados, {len(skipped)} pulados'
    })


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def list_deliveries(request):
    """
    Listar todas as entregas com filtros
    GET /api/deliveries/
    """
    from django.db.models import Q
    
    # Filtros
    search = request.GET.get('search', '').strip()
    company_id = request.GET.get('company', '').strip()
    status = request.GET.get('status', '').strip()
    
    # Query base - obter apenas a submission mais recente de cada obrigação
    from django.db.models import Max
    
    # Primeiro, encontrar a submission mais recente de cada obrigação
    latest_submissions = Submission.objects.values('obligation').annotate(
        latest_delivered_at=Max('delivered_at')
    )
    
    # Obter as submissions mais recentes
    latest_submission_ids = []
    for item in latest_submissions:
        latest_submission = Submission.objects.filter(
            obligation=item['obligation'],
            delivered_at=item['latest_delivered_at']
        ).first()
        if latest_submission:
            latest_submission_ids.append(latest_submission.id)
    
    # Query base com as submissions mais recentes
    qs = Submission.objects.filter(id__in=latest_submission_ids).select_related(
        'obligation__company', 
        'obligation__state', 
        'obligation__obligation_type',
        'delivered_by'
    ).prefetch_related('attachments').order_by('-delivered_at')
    
    # Aplicar filtros
    if search:
        qs = qs.filter(
            Q(obligation__company__name__icontains=search) |
            Q(obligation__obligation_name__icontains=search) |
            Q(obligation__obligation_type__name__icontains=search) |
            Q(comments__icontains=search)
        )
    
    if company_id:
        qs = qs.filter(obligation__company_id=company_id)
    
    if status:
        qs = qs.filter(submission_type=status)
    
    # Serializar dados
    deliveries = []
    for submission in qs:
        # Contar anexos
        attachments_count = 0
        if submission.receipt_file:
            attachments_count += 1
        attachments_count += submission.attachments.count()
        
        deliveries.append({
            'id': submission.id,
            'company': submission.obligation.company.name,
            'company_cnpj': submission.obligation.company.cnpj,
            'obligation': submission.obligation.obligation_name or submission.obligation.obligation_type.name,
            'competence': submission.obligation.competence,
            'state': submission.obligation.state.code,
            'delivery_date': submission.delivery_date.isoformat(),
            'delivered_at': submission.delivered_at.isoformat(),
            'delivered_by': submission.delivered_by.username if submission.delivered_by else None,
            'type': submission.submission_type,
            'comments': submission.comments,
            'attachments_count': attachments_count,
            'batch_id': str(submission.batch_id) if submission.batch_id else None,
            'has_receipt': bool(submission.receipt_file),
            'receipt_url': submission.receipt_file.url if submission.receipt_file else None
        })
    
    return Response({
        'deliveries': deliveries,
        'total': len(deliveries),
        'filters_applied': {
            'search': search,
            'company': company_id,
            'status': status
        }
    })
