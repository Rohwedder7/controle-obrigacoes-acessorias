from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.contrib.auth.models import User, Group
from django.db.models import Count, F, Q
from django.http import HttpResponse
import csv
from django.utils import timezone
from django.core.mail import send_mail
import io
import datetime
from openpyxl import Workbook

from .models import State, Company, ObligationType, Obligation, Submission, Notification
from .serializers import (
    UserSerializer, RegisterSerializer,
    StateSerializer, CompanySerializer, ObligationTypeSerializer,
    ObligationSerializer, SubmissionSerializer, NotificationSerializer
)
from .services import NotificationService, ObligationPlanningService
from .permissions import IsAdmin, IsUsuario, ReadOnlyOrCreateForUsuario, IsAdminOrReadOnly

class IsAuthenticatedOrCreate(permissions.IsAuthenticated):
    def has_permission(self, request, view):
        if view.action == 'create' and getattr(view, 'basename', '') == 'register':
            return True
        return super().has_permission(request, view)

@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def register(request):
    serializer = RegisterSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    user = serializer.save()
    return Response(UserSerializer(user).data, status=status.HTTP_201_CREATED)

@api_view(['GET'])
def list_users(request):
    """Lista todos os usuários do sistema"""
    users = User.objects.all()
    serializer = UserSerializer(users, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_current_user(request):
    """Retorna informações do usuário atual"""
    try:
        user = request.user
        groups = list(user.groups.values_list('name', flat=True))
        
        # Se for superuser, automaticamente tem role Admin
        if user.is_superuser and 'Admin' not in groups:
            admin_group, created = Group.objects.get_or_create(name='Admin')
            user.groups.add(admin_group)
            groups = list(user.groups.values_list('name', flat=True))
        
        return Response({
            'id': user.id,
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'groups': groups,
            'role': groups[0] if groups else 'Sem grupo',
            'is_superuser': user.is_superuser,
            'is_active': user.is_active,
            'is_admin': user.is_superuser or 'Admin' in groups
        })
    except Exception as e:
        return Response({
            'error': f'Erro interno do servidor: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class StateViewSet(viewsets.ModelViewSet):
    def get_permissions(self):
        # Usar nova permissão: leitura para autenticados, escrita apenas para Admin
        return [IsAdminOrReadOnly()]

    def perform_create(self, serializer):
        obj = serializer.save()
        audit(self.request.user, 'created', obj)

    def perform_update(self, serializer):
        obj = serializer.save()
        audit(self.request.user, 'updated', obj)

    def perform_destroy(self, instance):
        audit(self.request.user, 'deleted', instance)
        instance.delete()

    queryset = State.objects.all().order_by('code')
    serializer_class = StateSerializer
    permission_classes = [permissions.IsAuthenticated]

class CompanyViewSet(viewsets.ModelViewSet):
    def get_permissions(self):
        # Usar nova permissão: GET/POST para Admin/Usuario, PUT/PATCH/DELETE apenas Admin
        return [ReadOnlyOrCreateForUsuario()]

    def perform_create(self, serializer):
        obj = serializer.save()
        audit(self.request.user, 'created', obj)

    def perform_update(self, serializer):
        obj = serializer.save()
        audit(self.request.user, 'updated', obj)

    def perform_destroy(self, instance):
        audit(self.request.user, 'deleted', instance)
        instance.delete()

    queryset = Company.objects.all().order_by('name')
    serializer_class = CompanySerializer
    permission_classes = [permissions.IsAuthenticated]

class ObligationTypeViewSet(viewsets.ModelViewSet):
    def get_permissions(self):
        # Usar nova permissão: leitura para autenticados, escrita apenas para Admin
        return [IsAdminOrReadOnly()]

    def perform_create(self, serializer):
        obj = serializer.save()
        audit(self.request.user, 'created', obj)

    def perform_update(self, serializer):
        obj = serializer.save()
        audit(self.request.user, 'updated', obj)

    def perform_destroy(self, instance):
        audit(self.request.user, 'deleted', instance)
        instance.delete()

    queryset = ObligationType.objects.all().order_by('name')
    serializer_class = ObligationTypeSerializer
    permission_classes = [permissions.IsAuthenticated]

class ObligationViewSet(viewsets.ModelViewSet):
    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        company = request.query_params.get('company')
        state = request.query_params.get('state')
        otype = request.query_params.get('obligation_type')
        competence = request.query_params.get('competence')
        status = request.query_params.get('status')
        if company: qs = qs.filter(company__id=company)
        if state: qs = qs.filter(state__code=state)
        if otype: qs = qs.filter(obligation_type__id=otype)
        if competence: qs = qs.filter(competence=competence)
        if status == 'pendente':
            # Pendente = sem submissions aprovadas e não vencido
            qs = [o for o in qs if not o.submissions.filter(approval_status='approved').exists() and (not o.due_date or o.due_date >= timezone.now().date())]
        elif status == 'atrasado':
            # Atrasado = sem submissions aprovadas e vencido
            qs = [o for o in qs if not o.submissions.filter(approval_status='approved').exists() and o.due_date and o.due_date < timezone.now().date()]
        elif status == 'entregue':
            # Entregue = tem submissions aprovadas
            qs = [o for o in qs if o.submissions.filter(approval_status='approved').exists()]
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    def get_permissions(self):
        # Usar nova permissão: GET/POST para Admin/Usuario, PUT/PATCH/DELETE apenas Admin
        return [ReadOnlyOrCreateForUsuario()]

    def perform_create(self, serializer):
        obj = serializer.save()
        audit(self.request.user, 'created', obj)

    def perform_update(self, serializer):
        obj = serializer.save()
        audit(self.request.user, 'updated', obj)

    def perform_destroy(self, instance):
        audit(self.request.user, 'deleted', instance)
        instance.delete()

    queryset = Obligation.objects.select_related('company','state','obligation_type').prefetch_related('submissions').all().order_by('-due_date')
    serializer_class = ObligationSerializer
    permission_classes = [permissions.IsAuthenticated]

class SubmissionViewSet(viewsets.ModelViewSet):
    def get_permissions(self):
        # Usar nova permissão: GET/POST para Admin/Usuario, PUT/PATCH/DELETE apenas Admin
        return [ReadOnlyOrCreateForUsuario()]

    def perform_create(self, serializer):
        obj = serializer.save()
        audit(self.request.user, 'created', obj)

    def perform_update(self, serializer):
        obj = serializer.save()
        audit(self.request.user, 'updated', obj)

    def perform_destroy(self, instance):
        audit(self.request.user, 'deleted', instance)
        instance.delete()

    queryset = Submission.objects.select_related('obligation','delivered_by').all().order_by('-delivered_at')
    serializer_class = SubmissionSerializer
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [permissions.IsAuthenticated]

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def report_summary(request):
    # basic summary by company/state/obligation_type and status (delivered or not)
    data = []
    # counts of obligations and submissions
    qs = Obligation.objects.values(
        company=F('company__name'),
        state=F('state__code'),
        obligation_type=F('obligation_type__name'),
        competence=F('competence')
    ).annotate(total=Count('id'))
    delivered = Submission.objects.values('obligation').distinct().count()

    for row in qs:
        data.append(row)

    return Response({
        'summary': data,
        'total_obligations': Obligation.objects.count(),
        'total_submissions': Submission.objects.count(),
        'distinct_obligations_with_submission': delivered,
    })

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def report_detailed(request):
    """
    Relatório detalhado com filtros avançados
    """
    from django.db.models import Prefetch
    from datetime import datetime, date
    from calendar import monthrange
    
    # Parsear filtros
    company_ids = request.GET.getlist('company_id')
    obligation_name = request.GET.get('obligation_name', '').strip()
    obligation_type_ids = request.GET.getlist('obligation_type_id')
    competence_start = request.GET.get('competence_start', '').strip()
    competence_end = request.GET.get('competence_end', '').strip()
    due_start = request.GET.get('due_start', '').strip()
    due_end = request.GET.get('due_end', '').strip()
    status_filter = request.GET.get('status', '').strip()
    
    # Construir queryset base
    qs = Obligation.objects.select_related(
        'company', 'state', 'obligation_type', 'created_by'
    ).prefetch_related(
        Prefetch('submissions', queryset=Submission.objects.select_related('delivered_by').order_by('-delivered_at'))
    )
    
    # Aplicar filtros
    if company_ids:
        qs = qs.filter(company_id__in=company_ids)
    
    if obligation_name:
        qs = qs.filter(
            Q(obligation_name__icontains=obligation_name) | 
            Q(obligation_type__name__icontains=obligation_name)
        )
    
    if obligation_type_ids:
        qs = qs.filter(obligation_type_id__in=obligation_type_ids)
    
    # Filtros de competência
    if competence_start:
        try:
            month, year = competence_start.split('/')
            start_date = date(int(year), int(month), 1)
            qs = qs.filter(due_date__gte=start_date)
        except (ValueError, IndexError):
            pass
    
    if competence_end:
        try:
            month, year = competence_end.split('/')
            last_day = monthrange(int(year), int(month))[1]
            end_date = date(int(year), int(month), last_day)
            qs = qs.filter(due_date__lte=end_date)
        except (ValueError, IndexError):
            pass
    
    # Filtros de data de vencimento
    if due_start:
        try:
            start_date = datetime.strptime(due_start, '%Y-%m-%d').date()
            qs = qs.filter(due_date__gte=start_date)
        except ValueError:
            pass
    
    if due_end:
        try:
            end_date = datetime.strptime(due_end, '%Y-%m-%d').date()
            qs = qs.filter(due_date__lte=end_date)
        except ValueError:
            pass
    
    # Ordenação padrão
    qs = qs.order_by('company__name', 'obligation_type__name', '-competence')
    
    # Calcular totais
    total_obligations = qs.count()
    today = timezone.now().date()
    
    # Status das obrigações - considerar apenas aprovadas
    delivered_count = 0
    pending_count = 0
    late_count = 0
    
    rows = []
    for obligation in qs:
        # Determinar status - considerar apenas submissions aprovadas
        latest_submission = obligation.submissions.filter(approval_status='approved').order_by('-delivered_at').first()
        
        if latest_submission:
            status = 'entregue'
            delivered_count += 1
            days_late = 0
        elif obligation.due_date < today:
            status = 'atrasado'
            late_count += 1
            days_late = (today - obligation.due_date).days
        else:
            status = 'pendente'
            pending_count += 1
            days_late = 0
        
        # Aplicar filtro de status se especificado
        if status_filter and status != status_filter:
            continue
        
        # Contar anexos (receipt_file + attachments)
        receipt_files = obligation.submissions.filter(receipt_file__isnull=False).count()
        attachment_files = sum(submission.attachments.count() for submission in obligation.submissions.all())
        total_attachments = receipt_files + attachment_files
        
        # Informações da última entrega
        submission_info = None
        if latest_submission:
            submission_info = {
                'delivered_at': latest_submission.delivered_at.isoformat(),
                'delivered_by': latest_submission.delivered_by.username if latest_submission.delivered_by else None,
                'submission_type': latest_submission.submission_type,
                'delivery_date': latest_submission.delivery_date.isoformat(),
                'comments': latest_submission.comments or '',
                'has_receipt_file': bool(latest_submission.receipt_file),
                'attachments_count': latest_submission.attachments.count()
            }
        
        row = {
            'company': obligation.company.name,
            'cnpj': obligation.company.cnpj or '',
            'state': obligation.state.code,
            'obligation_type': obligation.obligation_type.name,
            'obligation_name': obligation.obligation_name or '',
            'competence': obligation.competence,
            'due_date': obligation.due_date.isoformat(),
            'delivery_deadline': obligation.delivery_deadline.isoformat() if obligation.delivery_deadline else None,
            'status': status,
            'days_late': days_late,
            'notes': obligation.notes or '',
            'total_attachments': total_attachments,
            'submission_info': submission_info,
            'created_by': obligation.created_by.username if obligation.created_by else None,
            'created_at': obligation.created_at.isoformat()
        }
        rows.append(row)
    
    # Agregações por empresa
    by_company = []
    company_stats = {}
    for row in rows:
        company = row['company']
        if company not in company_stats:
            company_stats[company] = {'count': 0, 'pending': 0, 'late': 0, 'delivered': 0}
        
        company_stats[company]['count'] += 1
        if row['status'] == 'pendente':
            company_stats[company]['pending'] += 1
        elif row['status'] == 'atrasado':
            company_stats[company]['late'] += 1
        elif row['status'] == 'entregue':
            company_stats[company]['delivered'] += 1
    
    for company, stats in company_stats.items():
        by_company.append({
            'company': company,
            'count': stats['count'],
            'pending': stats['pending'],
            'late': stats['late'],
            'delivered': stats['delivered']
        })
    
    # Agregações por estado
    by_state = []
    state_stats = {}
    for row in rows:
        state = row['state']
        if state not in state_stats:
            state_stats[state] = {'count': 0, 'pending': 0, 'late': 0, 'delivered': 0}
        
        state_stats[state]['count'] += 1
        if row['status'] == 'pendente':
            state_stats[state]['pending'] += 1
        elif row['status'] == 'atrasado':
            state_stats[state]['late'] += 1
        elif row['status'] == 'entregue':
            state_stats[state]['delivered'] += 1
    
    for state, stats in state_stats.items():
        by_state.append({
            'state': state,
            'count': stats['count'],
            'pending': stats['pending'],
            'late': stats['late'],
            'delivered': stats['delivered']
        })
    
    # Agregações por tipo
    by_type = []
    type_stats = {}
    for row in rows:
        obligation_type = row['obligation_type']
        if obligation_type not in type_stats:
            type_stats[obligation_type] = {'count': 0, 'pending': 0, 'late': 0, 'delivered': 0}
        
        type_stats[obligation_type]['count'] += 1
        if row['status'] == 'pendente':
            type_stats[obligation_type]['pending'] += 1
        elif row['status'] == 'atrasado':
            type_stats[obligation_type]['late'] += 1
        elif row['status'] == 'entregue':
            type_stats[obligation_type]['delivered'] += 1
    
    for obligation_type, stats in type_stats.items():
        by_type.append({
            'type': obligation_type,
            'count': stats['count'],
            'pending': stats['pending'],
            'late': stats['late'],
            'delivered': stats['delivered']
        })
    
    return Response({
        'filters_applied': {
            'company_ids': company_ids,
            'obligation_name': obligation_name,
            'obligation_type_ids': obligation_type_ids,
            'competence_start': competence_start,
            'competence_end': competence_end,
            'due_start': due_start,
            'due_end': due_end,
            'status': status_filter
        },
        'totals': {
            'obligations': len(rows),
            'submissions': delivered_count,
            'pending': pending_count,
            'late': late_count,
            'delivered': delivered_count
        },
        'by_company': sorted(by_company, key=lambda x: x['count'], reverse=True),
        'by_state': sorted(by_state, key=lambda x: x['count'], reverse=True),
        'by_type': sorted(by_type, key=lambda x: x['count'], reverse=True),
        'rows': rows
    })

def _apply_report_filters(request):
    """
    Função auxiliar para aplicar filtros de relatório
    """
    from datetime import datetime, date
    from calendar import monthrange
    
    # Parsear filtros
    company_ids = request.GET.getlist('company_id')
    obligation_name = request.GET.get('obligation_name', '').strip()
    obligation_type_ids = request.GET.getlist('obligation_type_id')
    competence_start = request.GET.get('competence_start', '').strip()
    competence_end = request.GET.get('competence_end', '').strip()
    due_start = request.GET.get('due_start', '').strip()
    due_end = request.GET.get('due_end', '').strip()
    status_filter = request.GET.get('status', '').strip()
    
    # Construir queryset base
    qs = Obligation.objects.select_related(
        'company', 'state', 'obligation_type', 'created_by'
    ).prefetch_related('submissions')
    
    # Aplicar filtros
    if company_ids:
        qs = qs.filter(company_id__in=company_ids)
    
    if obligation_name:
        qs = qs.filter(
            Q(obligation_name__icontains=obligation_name) | 
            Q(obligation_type__name__icontains=obligation_name)
        )
    
    if obligation_type_ids:
        qs = qs.filter(obligation_type_id__in=obligation_type_ids)
    
    # Filtros de competência
    if competence_start:
        try:
            month, year = competence_start.split('/')
            start_date = date(int(year), int(month), 1)
            qs = qs.filter(due_date__gte=start_date)
        except (ValueError, IndexError):
            pass
    
    if competence_end:
        try:
            month, year = competence_end.split('/')
            last_day = monthrange(int(year), int(month))[1]
            end_date = date(int(year), int(month), last_day)
            qs = qs.filter(due_date__lte=end_date)
        except (ValueError, IndexError):
            pass
    
    # Filtros de data de vencimento
    if due_start:
        try:
            start_date = datetime.strptime(due_start, '%Y-%m-%d').date()
            qs = qs.filter(due_date__gte=start_date)
        except ValueError:
            pass
    
    if due_end:
        try:
            end_date = datetime.strptime(due_end, '%Y-%m-%d').date()
            qs = qs.filter(due_date__lte=end_date)
        except ValueError:
            pass
    
    # Ordenação padrão
    qs = qs.order_by('company__name', 'obligation_type__name', '-competence')
    
    return qs, status_filter

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def report_csv(request):
    # Export obligations + latest submission (if any) with filters
    qs, status_filter = _apply_report_filters(request)
    
    # Gerar nome do arquivo com filtros
    filename = "obrigacoes.csv"
    if status_filter:
        filename = f"obrigacoes_{status_filter}.csv"
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    
    # Cabeçalhos expandidos
    writer.writerow([
        'Empresa', 'CNPJ', 'Estado', 'Tipo de Obrigação', 'Nome da Obrigação', 
        'Competência', 'Vencimento', 'Prazo Entrega', 'Status', 'Entregue em', 
        'Entregue por', 'Dias de Atraso', 'Anexos', 'Criado por', 'Criado em', 'Notas'
    ])
    
    today = timezone.now().date()
    for o in qs:
        # Considerar apenas submissions aprovadas
        sub = o.submissions.filter(approval_status='approved').order_by('-delivered_at').first()
        
        # Determinar status
        if sub:
            status = 'entregue'
            days_late = 0
        elif o.due_date < today:
            status = 'atrasado'
            days_late = (today - o.due_date).days
        else:
            status = 'pendente'
            days_late = 0
        
        # Aplicar filtro de status se especificado
        if status_filter and status != status_filter:
            continue
        
        # Contar anexos
        attachments_count = o.submissions.filter(receipt_file__isnull=False).count()
        
        writer.writerow([
            o.company.name,
            o.company.cnpj or '',
            o.state.code,
            o.obligation_type.name,
            o.obligation_name or '',
            o.competence,
            o.due_date.isoformat(),
            o.delivery_deadline.isoformat() if o.delivery_deadline else '',
            status,
            sub.delivery_date.isoformat() if sub else '',
            sub.delivered_by.username if sub and sub.delivered_by else '',
            days_late,
            attachments_count,
            o.created_by.username if o.created_by else '',
            o.created_at.isoformat(),
            o.notes or ''
        ])
    return response

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def report_xlsx(request):
    # Export obligations + latest submission (if any) with filters
    qs, status_filter = _apply_report_filters(request)
    
    # Gerar nome do arquivo com filtros
    filename = "obrigacoes.xlsx"
    if status_filter:
        filename = f"obrigacoes_{status_filter}.xlsx"
    
    wb = Workbook()
    
    # Aba principal - Obrigações
    ws = wb.active
    ws.title = "Obrigações"
    
    # Cabeçalhos expandidos
    headers = [
        'Empresa', 'CNPJ', 'Estado', 'Tipo de Obrigação', 'Nome da Obrigação', 
        'Competência', 'Vencimento', 'Prazo Entrega', 'Status', 'Entregue em', 
        'Entregue por', 'Dias de Atraso', 'Anexos', 'Criado por', 'Criado em', 'Notas'
    ]
    ws.append(headers)
    
    # Aplicar formatação ao cabeçalho
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Congelar primeira linha
    ws.freeze_panes = "A2"
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:{chr(65 + len(headers) - 1)}1"
    
    # Ajustar largura das colunas
    column_widths = [20, 15, 8, 25, 25, 12, 12, 12, 10, 12, 15, 10, 8, 15, 12, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    today = timezone.now().date()
    row_num = 2
    
    for o in qs:
        # Considerar apenas submissions aprovadas
        sub = o.submissions.filter(approval_status='approved').order_by('-delivered_at').first()
        
        # Determinar status
        if sub:
            status = 'entregue'
            days_late = 0
        elif o.due_date < today:
            status = 'atrasado'
            days_late = (today - o.due_date).days
        else:
            status = 'pendente'
            days_late = 0
        
        # Aplicar filtro de status se especificado
        if status_filter and status != status_filter:
            continue
        
        # Contar anexos
        attachments_count = o.submissions.filter(receipt_file__isnull=False).count()
        
        row_data = [
            o.company.name,
            o.company.cnpj or '',
            o.state.code,
            o.obligation_type.name,
            o.obligation_name or '',
            o.competence,
            o.due_date.isoformat(),
            o.delivery_deadline.isoformat() if o.delivery_deadline else '',
            status,
            sub.delivery_date.isoformat() if sub else '',
            sub.delivered_by.username if sub and sub.delivered_by else '',
            days_late,
            attachments_count,
            o.created_by.username if o.created_by else '',
            o.created_at.isoformat(),
            o.notes or ''
        ]
        ws.append(row_data)
        row_num += 1
    
    # Aba de Resumo
    ws_summary = wb.create_sheet("Resumo")
    
    # Calcular totais
    total_obligations = row_num - 2  # -2 para excluir cabeçalho e linha vazia
    delivered_count = sum(1 for row in ws.iter_rows(min_row=2, max_row=row_num) if row[8].value == 'entregue')
    pending_count = sum(1 for row in ws.iter_rows(min_row=2, max_row=row_num) if row[8].value == 'pendente')
    late_count = sum(1 for row in ws.iter_rows(min_row=2, max_row=row_num) if row[8].value == 'atrasado')
    
    # Adicionar resumo
    ws_summary.append(['Resumo de Obrigações'])
    ws_summary.append(['Total de Obrigações', total_obligations])
    ws_summary.append(['Entregues', delivered_count])
    ws_summary.append(['Pendentes', pending_count])
    ws_summary.append(['Atrasadas', late_count])
    ws_summary.append([])
    ws_summary.append(['Filtros Aplicados'])
    ws_summary.append(['Status', status_filter or 'Todos'])
    
    # Formatação da aba resumo
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A7'].font = Font(bold=True, size=12)
    
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp

ROLE_ADMIN = 'Admin'
ROLE_FISCAL = 'Fiscal'
ROLE_VISUALIZADOR = 'Visualizador'

def user_has_role(user, role):
    return user.is_superuser or user.groups.filter(name=role).exists()

def audit(user, action, instance, changes=None):
    from .models import AuditLog
    try:
        AuditLog.objects.create(
            user=user if user.is_authenticated else None,
            action=action,
            model=instance.__class__.__name__,
            object_id=str(instance.pk),
            changes=changes or {},
        )
    except Exception:
        pass


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def dashboard_metrics(request):
    from datetime import date, timedelta, datetime
    from django.db.models import Count, Q
    from django.db.models.functions import TruncMonth
    
    today = timezone.now().date()
    current_month = today.strftime('%m/%Y')
    
    # Parâmetros de filtro de data
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Query base para obrigações
    obligations_query = Obligation.objects.all()

    # Aplicar filtros de data se fornecidos
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            obligations_query = obligations_query.filter(due_date__gte=start_date_obj)
        except ValueError:
            pass

    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            obligations_query = obligations_query.filter(due_date__lte=end_date_obj)
        except ValueError:
            pass
    
    # Métricas gerais - considerar apenas submissions aprovadas
    total_obligations = obligations_query.count()
    pending_obligations = obligations_query.exclude(submissions__approval_status='approved').distinct().count()
    delivered_obligations = obligations_query.filter(submissions__approval_status='approved').distinct().count()
    overdue_obligations = obligations_query.filter(
        due_date__lt=today
    ).exclude(submissions__approval_status='approved').distinct().count()
    
    # Taxa de cumprimento geral
    compliance_rate = (delivered_obligations / total_obligations * 100) if total_obligations > 0 else 0
    
    # Performance por usuário responsável (top 5) - considerar apenas aprovadas
    user_performance = obligations_query.values(
        'responsible_user__username', 
        'responsible_user__first_name', 
        'responsible_user__last_name'
    ).annotate(
        total=Count('id'),
        delivered=Count('id', filter=Q(submissions__approval_status='approved')),
        pending=Count('id', filter=~Q(submissions__approval_status='approved')),
        overdue=Count('id', filter=Q(due_date__lt=today) & ~Q(submissions__approval_status='approved'))
    ).order_by('-total')[:5]
    
    # Top 5 empresas com mais obrigações - considerar apenas aprovadas
    company_performance = obligations_query.values('company__name', 'company__cnpj').annotate(
        total=Count('id'),
        delivered=Count('id', filter=Q(submissions__approval_status='approved')),
        pending=Count('id', filter=~Q(submissions__approval_status='approved')),
        overdue=Count('id', filter=Q(due_date__lt=today) & ~Q(submissions__approval_status='approved'))
    ).order_by('-total')[:5]
    
    # Obrigações por tipo (top 5) - considerar apenas aprovadas
    obligation_type_performance = obligations_query.values('obligation_type__name').annotate(
        total=Count('id'),
        delivered=Count('id', filter=Q(submissions__approval_status='approved')),
        pending=Count('id', filter=~Q(submissions__approval_status='approved')),
        overdue=Count('id', filter=Q(due_date__lt=today) & ~Q(submissions__approval_status='approved'))
    ).order_by('-total')[:5]
    
    # Tendência mensal (últimos 6 meses) - considerar apenas aprovadas
    monthly_trend = obligations_query.annotate(
        month=TruncMonth('due_date')
    ).values('month').annotate(
        total=Count('id'),
        delivered=Count('id', filter=Q(submissions__approval_status='approved')),
        pending=Count('id', filter=~Q(submissions__approval_status='approved')),
        overdue=Count('id', filter=Q(due_date__lt=today) & ~Q(submissions__approval_status='approved'))
    ).order_by('-month')[:6]
    
    # Obrigações que vencem nos próximos 7 dias (sem aprovação)
    upcoming_obligations = obligations_query.filter(
        due_date__lte=today + timedelta(days=7),
        due_date__gte=today
    ).exclude(
        submissions__approval_status='approved'
    ).select_related('company', 'obligation_type', 'state', 'responsible_user')[:10]
    
    upcoming_serialized = []
    for obligation in upcoming_obligations:
        upcoming_serialized.append({
            'id': obligation.id,
            'company_name': obligation.company.name,
            'obligation_name': obligation.obligation_name or obligation.obligation_type.name,
            'obligation_type': obligation.obligation_type.name,
            'state': obligation.state.code,
            'competence': obligation.competence,
            'due_date': obligation.due_date,
            'responsible_user': obligation.responsible_user.username if obligation.responsible_user else 'Sem responsável',
            'responsible_user_full_name': f"{obligation.responsible_user.first_name} {obligation.responsible_user.last_name}".strip() if obligation.responsible_user else 'Sem responsável',
            'days_until_due': (obligation.due_date - today).days
        })
    
    # Obrigações críticas (em atraso há mais de 5 dias sem aprovação)
    critical_obligations = obligations_query.filter(
        due_date__lt=today - timedelta(days=5)
    ).exclude(
        submissions__approval_status='approved'
    ).select_related('company', 'obligation_type', 'state', 'responsible_user')[:10]
    
    critical_serialized = []
    for obligation in critical_obligations:
        critical_serialized.append({
            'id': obligation.id,
            'company_name': obligation.company.name,
            'obligation_name': obligation.obligation_name or obligation.obligation_type.name,
            'obligation_type': obligation.obligation_type.name,
            'state': obligation.state.code,
            'competence': obligation.competence,
            'due_date': obligation.due_date,
            'responsible_user': obligation.responsible_user.username if obligation.responsible_user else 'Sem responsável',
            'responsible_user_full_name': f"{obligation.responsible_user.first_name} {obligation.responsible_user.last_name}".strip() if obligation.responsible_user else 'Sem responsável',
            'days_overdue': (today - obligation.due_date).days
        })
    
    # Empresas com obrigações do mês atual (mantendo compatibilidade) - considerar apenas aprovadas
    companies_data = []
    for company in Company.objects.filter(active=True).prefetch_related('obligations'):
        current_obligations = company.obligations.filter(competence=current_month)
        pending = sum(1 for o in current_obligations if not o.submissions.filter(approval_status='approved').exists())
        delivered = sum(1 for o in current_obligations if o.submissions.filter(approval_status='approved').exists())
        
        companies_data.append({
            'id': company.id,
            'name': company.name,
            'cnpj': company.cnpj,
            'pending_count': pending,
            'delivered_count': delivered,
            'total_count': current_obligations.count()
        })
    
    # Dados para gráficos (mantendo compatibilidade) - considerar apenas aprovadas
    from collections import defaultdict
    atrasos = defaultdict(int)
    total = defaultdict(int)
    
    for o in Obligation.objects.all():
        ym = (o.due_date or today).strftime('%Y-%m')
        total[ym] += 1
        if not o.submissions.filter(approval_status='approved').exists() and o.due_date and o.due_date < today:
            atrasos[ym] += 1
    
    months = sorted(total.keys())[-6:]
    atraso_series = [atrasos.get(m,0) for m in months]
    
    return Response({
        # Métricas gerais
        'total_obligations': total_obligations,
        'pending_obligations': pending_obligations,
        'delivered_obligations': delivered_obligations,
        'overdue_obligations': overdue_obligations,
        'compliance_rate': round(compliance_rate, 2),
        
        # Performance por usuário
        'user_performance': list(user_performance),
        
        # Performance por empresa
        'company_performance': list(company_performance),
        
        # Performance por tipo de obrigação
        'obligation_type_performance': list(obligation_type_performance),
        
        # Tendência mensal
        'monthly_trend': list(monthly_trend),
        
        # Obrigações próximas do vencimento
        'upcoming_obligations': upcoming_serialized,
        
        # Obrigações críticas
        'critical_obligations': critical_serialized,
        
        # Dados para gráficos (compatibilidade)
        'months': months,
        'late_counts': atraso_series,
        'current_month': current_month,
        'companies': companies_data,
        
        # Métricas do mês atual (compatibilidade) - considerar apenas aprovadas
        'total_companies': Company.objects.filter(active=True).count(),
        'total_obligations_month': Obligation.objects.filter(competence=current_month).count(),
        'pending_obligations_month': sum(1 for o in Obligation.objects.filter(competence=current_month) if not o.submissions.filter(approval_status='approved').exists()),
        'delivered_obligations_month': sum(1 for o in Obligation.objects.filter(competence=current_month) if o.submissions.filter(approval_status='approved').exists()),
        
        # Estatísticas de notificações
        'notifications': {
            'total': Notification.objects.count(),
            'unread': Notification.objects.filter(is_read=False).count()
        },
        
        # Filtros aplicados
        'filters_applied': {
            'start_date': start_date,
            'end_date': end_date
        }
    })

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def bulk_import_obligations(request):
    """Upload em massa de obrigações via planilha Excel"""
    from openpyxl import load_workbook
    file = request.FILES.get('file')
    if not file:
        return Response({'detail':'Arquivo não enviado.'}, status=400)
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        created = 0
        errors = []
        
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:  # header
                continue
                
            try:
                # Mapear colunas da planilha
                company_name, state_code, otype_name, obligation_name, competence, due_date, delivery_deadline, responsible_username, validity_start, validity_end, notes = row[:11]
                
                if not company_name or not state_code or not otype_name or not competence or not due_date:
                    errors.append(f"Linha {i}: Campos obrigatórios não preenchidos")
                    continue
                
                # Converter competência para formato MM/YYYY
                competence_formatted = None
                try:
                    if isinstance(competence, datetime.datetime):
                        # Se for datetime, converter para MM/YYYY
                        competence_formatted = competence.strftime('%m/%Y')
                    elif isinstance(competence, datetime.date):
                        # Se for date, converter para MM/YYYY
                        competence_formatted = competence.strftime('%m/%Y')
                    elif isinstance(competence, str):
                        # Se for string, tentar diferentes formatos
                        competence_str = str(competence).strip()
                        if '/' in competence_str:
                            # Já está no formato MM/YYYY
                            competence_formatted = competence_str
                        elif '-' in competence_str:
                            # Formato YYYY-MM-DD ou similar
                            try:
                                if len(competence_str) >= 7:  # YYYY-MM
                                    year, month = competence_str[:7].split('-')
                                    competence_formatted = f"{month.zfill(2)}/{year}"
                                else:
                                    competence_formatted = competence_str
                            except:
                                competence_formatted = competence_str
                        else:
                            competence_formatted = competence_str
                    else:
                        competence_formatted = str(competence)
                except Exception as e:
                    errors.append(f"Linha {i}: Erro ao converter competência '{competence}': {str(e)}")
                    continue
                
                # Buscar ou criar empresa
                company = Company.objects.filter(name=str(company_name)).first()
                if not company:
                    errors.append(f"Linha {i}: Empresa '{company_name}' não encontrada")
                    continue
                
                # Buscar estado
                state = State.objects.filter(code=str(state_code).upper()).first()
                if not state:
                    errors.append(f"Linha {i}: Estado '{state_code}' não encontrado")
                    continue
                
                # Buscar ou criar tipo de obrigação
                otype, _ = ObligationType.objects.get_or_create(name=str(otype_name))
                
                # Buscar usuário responsável se informado
                responsible_user = None
                if responsible_username:
                    try:
                        responsible_user = User.objects.get(username=str(responsible_username))
                    except User.DoesNotExist:
                        errors.append(f"Linha {i}: Usuário responsável '{responsible_username}' não encontrado")
                        continue
                
                # Criar obrigação
                obj, made = Obligation.objects.get_or_create(
                    company=company, 
                    state=state, 
                    obligation_type=otype, 
                    competence=competence_formatted,
                    defaults={
                        'obligation_name': obligation_name or '',
                        'due_date': due_date,
                        'delivery_deadline': delivery_deadline,
                        'responsible_user': responsible_user,
                        'validity_start_date': validity_start or '2024-01-01',
                        'validity_end_date': validity_end or '2024-12-31',
                        'created_by': request.user if request.user.is_authenticated else None,
                        'notes': notes or ''
                    }
                )
                if made: 
                    created += 1
                    
            except Exception as e:
                errors.append(f"Linha {i}: Erro - {str(e)}")
        
        return Response({
            'created': created,
            'errors': errors,
            'total_processed': i - 1
        })
        
    except Exception as e:
        return Response({'detail': f'Erro ao processar arquivo: {str(e)}'}, status=400)

@api_view(['POST'])
def bulk_import_companies(request):
    """Upload em massa de empresas via planilha Excel"""
    from openpyxl import load_workbook
    file = request.FILES.get('file')
    if not file:
        return Response({'detail':'Arquivo não enviado.'}, status=400)
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        created = 0
        errors = []
        
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:  # header
                continue
                
            try:
                # Mapear colunas da planilha
                name, cnpj, fantasy_name, email, phone, address, responsible = row[:7]
                
                if not name:
                    errors.append(f"Linha {i}: Nome da empresa é obrigatório")
                    continue
                
                # Criar empresa
                obj, made = Company.objects.get_or_create(
                    name=str(name),
                    defaults={
                        'cnpj': str(cnpj) if cnpj else '',
                        'fantasy_name': str(fantasy_name) if fantasy_name else '',
                        'email': str(email) if email else '',
                        'phone': str(phone) if phone else '',
                        'address': str(address) if address else '',
                        'responsible': str(responsible) if responsible else ''
                    }
                )
                if made: 
                    created += 1
                    
            except Exception as e:
                errors.append(f"Linha {i}: Erro - {str(e)}")
        
        return Response({
            'created': created,
            'errors': errors,
            'total_processed': i - 1
        })
        
    except Exception as e:
        return Response({'detail': f'Erro ao processar arquivo: {str(e)}'}, status=400)

@api_view(['GET'])
def download_template(request, template_type):
    """Download de templates de planilha"""
    import os
    from django.http import FileResponse
    from django.conf import settings
    
    template_files = {
        'companies': 'template_empresas.xlsx',
        'obligations': 'template_obrigacoes.xlsx'
    }
    
    if template_type not in template_files:
        return Response({'detail': 'Tipo de template inválido'}, status=400)
    
    file_path = os.path.join(settings.MEDIA_ROOT, 'templates', template_files[template_type])
    
    if not os.path.exists(file_path):
        return Response({'detail': 'Template não encontrado'}, status=404)
    
    return FileResponse(
        open(file_path, 'rb'),
        as_attachment=True,
        filename=template_files[template_type]
    )

@api_view(['POST'])
def send_reminders(request):
    days = int(request.data.get('days', 3))
    today = timezone.now().date()
    target = today + datetime.timedelta(days=days)
    count = 0
    recipients = [u.email for u in User.objects.exclude(email='').all()]
    for o in Obligation.objects.filter(due_date=target):
        try:
            send_mail(
                subject=f'[Alerta] Obrigação vence em {days} dia(s): {o.obligation_type.name} — {o.company.name}/{o.state.code} ({o.competence})',
                message=f'Vencimento: {o.due_date}\nPrazo de entrega: {o.delivery_deadline or "-"}\nNotas: {o.notes or "-"}',
                from_email=EMAIL_HOST_USER if 'EMAIL_HOST_USER' in globals() else None,
                recipient_list=recipients,
                fail_silently=True
            )
            count += 1
        except Exception:
            pass
    return Response({'emails_sent_for': count})

# Views para Notificações
@api_view(['GET'])
def get_notifications(request):
    """Lista notificações do usuário"""
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    serializer = NotificationSerializer(notifications, many=True)
    return Response(serializer.data)

@api_view(['POST'])
def mark_notification_read(request, notification_id):
    """Marca uma notificação como lida"""
    try:
        notification = Notification.objects.get(id=notification_id, user=request.user)
        notification.is_read = True
        notification.read_at = timezone.now()
        notification.save()
        return Response({'status': 'success'})
    except Notification.DoesNotExist:
        return Response({'error': 'Notificação não encontrada'}, status=404)

@api_view(['POST'])
def mark_all_notifications_read(request):
    """Marca todas as notificações como lidas"""
    Notification.objects.filter(user=request.user, is_read=False).update(
        is_read=True, read_at=timezone.now()
    )
    return Response({'status': 'success'})

@api_view(['GET'])
def get_notification_stats(request):
    """Estatísticas de notificações"""
    total = Notification.objects.filter(user=request.user).count()
    unread = Notification.objects.filter(user=request.user, is_read=False).count()
    
    return Response({
        'total': total,
        'unread': unread,
        'read': total - unread
    })

# Views para Planejamento Automático
@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def generate_obligations(request):
    """Gera obrigações automaticamente"""
    months_ahead = request.data.get('months_ahead', 3)
    company_id = request.data.get('company_id')
    obligation_type_id = request.data.get('obligation_type_id')
    obligation_name = request.data.get('obligation_name')
    state_id = request.data.get('state_id')
    
    try:
        generated, skipped = ObligationPlanningService.generate_obligations_for_period(
            months_ahead, company_id, obligation_type_id, obligation_name, state_id
        )
        
        return Response({
            'generated': generated,
            'skipped': skipped,
            'message': f'{generated} obrigações geradas, {skipped} já existiam'
        })
    except Exception as e:
        return Response({'error': str(e)}, status=400)

# Views para Sistema de Alertas
@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def check_due_dates(request):
    """Verifica obrigações que vencem em breve"""
    days_ahead = request.data.get('days_ahead', 7)
    
    try:
        notifications_created = NotificationService.check_due_dates(days_ahead)
        return Response({
            'notifications_created': notifications_created,
            'message': f'{notifications_created} notificações criadas'
        })
    except Exception as e:
        return Response({'error': str(e)}, status=400)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def check_overdue_obligations(request):
    """Verifica obrigações em atraso"""
    try:
        notifications_created = NotificationService.check_overdue_obligations()
        return Response({
            'notifications_created': notifications_created,
            'message': f'{notifications_created} notificações de atraso criadas'
        })
    except Exception as e:
        return Response({'error': str(e)}, status=400)

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def send_email_notifications(request):
    """Envia notificações por email"""
    try:
        emails_sent = NotificationService.send_email_notifications()
        return Response({
            'emails_sent': emails_sent,
            'message': f'{emails_sent} emails enviados'
        })
    except Exception as e:
        return Response({'error': str(e)}, status=400)

# Views para Relatórios Avançados
@api_view(['GET'])
def advanced_reports_summary(request):
    """Relatório avançado com análises detalhadas"""
    from datetime import date
    from django.db.models import Count, Q
    
    # Parâmetros de filtro
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    user_id = request.GET.get('user_id')
    company_id = request.GET.get('company_id')
    state_id = request.GET.get('state_id')
    
    # Query base
    obligations = Obligation.objects.all()
    
    # Aplicar filtros
    if start_date:
        obligations = obligations.filter(validity_start_date__gte=start_date)
    if end_date:
        obligations = obligations.filter(validity_end_date__lte=end_date)
    if user_id:
        obligations = obligations.filter(responsible_user_id=user_id)
    if company_id:
        obligations = obligations.filter(company_id=company_id)
    if state_id:
        obligations = obligations.filter(state_id=state_id)
    
    # Estatísticas gerais
    total_obligations = obligations.count()
    delivered_obligations = obligations.filter(submissions__isnull=False).distinct().count()
    pending_obligations = total_obligations - delivered_obligations
    
    # Obrigações em atraso
    today = date.today()
    overdue_obligations = obligations.filter(
        due_date__lt=today,
        submissions__isnull=True
    ).count()
    
    # Taxa de cumprimento
    compliance_rate = (delivered_obligations / total_obligations * 100) if total_obligations > 0 else 0
    
    # Análise por usuário responsável
    user_stats = obligations.values('responsible_user__username', 'responsible_user__first_name', 'responsible_user__last_name').annotate(
        total=Count('id'),
        delivered=Count('id', filter=Q(submissions__isnull=False)),
        pending=Count('id', filter=Q(submissions__isnull=True)),
        overdue=Count('id', filter=Q(due_date__lt=today, submissions__isnull=True))
    ).order_by('-total')
    
    # Análise por empresa
    company_stats = obligations.values('company__name', 'company__cnpj').annotate(
        total=Count('id'),
        delivered=Count('id', filter=Q(submissions__isnull=False)),
        pending=Count('id', filter=Q(submissions__isnull=True)),
        overdue=Count('id', filter=Q(due_date__lt=today, submissions__isnull=True))
    ).order_by('-total')
    
    # Análise por estado
    state_stats = obligations.values('state__code', 'state__name').annotate(
        total=Count('id'),
        delivered=Count('id', filter=Q(submissions__isnull=False)),
        pending=Count('id', filter=Q(submissions__isnull=True)),
        overdue=Count('id', filter=Q(due_date__lt=today, submissions__isnull=True))
    ).order_by('-total')
    
    # Análise por tipo de obrigação
    obligation_type_stats = obligations.values('obligation_type__name', 'obligation_type__recurrence').annotate(
        total=Count('id'),
        delivered=Count('id', filter=Q(submissions__isnull=False)),
        pending=Count('id', filter=Q(submissions__isnull=True)),
        overdue=Count('id', filter=Q(due_date__lt=today, submissions__isnull=True))
    ).order_by('-total')
    
    # Análise temporal (últimos 12 meses)
    from django.db.models.functions import TruncMonth
    monthly_stats = obligations.annotate(
        month=TruncMonth('due_date')
    ).values('month').annotate(
        total=Count('id'),
        delivered=Count('id', filter=Q(submissions__isnull=False)),
        pending=Count('id', filter=Q(submissions__isnull=True)),
        overdue=Count('id', filter=Q(due_date__lt=today, submissions__isnull=True))
    ).order_by('month')[:12]
    
    return Response({
        'summary': {
            'total_obligations': total_obligations,
            'delivered_obligations': delivered_obligations,
            'pending_obligations': pending_obligations,
            'overdue_obligations': overdue_obligations,
            'compliance_rate': round(compliance_rate, 2)
        },
        'by_user': list(user_stats),
        'by_company': list(company_stats),
        'by_state': list(state_stats),
        'by_obligation_type': list(obligation_type_stats),
        'monthly_trend': list(monthly_stats),
        'filters_applied': {
            'start_date': start_date,
            'end_date': end_date,
            'user_id': user_id,
            'company_id': company_id,
            'state_id': state_id
        }
    })

@api_view(['GET'])
def user_performance_report(request, user_id):
    """Relatório de performance de um usuário específico"""
    from datetime import date
    from django.db.models import Avg, Count, Q
    
    user = User.objects.get(id=user_id)
    
    # Obrigações do usuário
    user_obligations = Obligation.objects.filter(responsible_user=user)
    
    # Estatísticas básicas
    total = user_obligations.count()
    delivered = user_obligations.filter(submissions__isnull=False).distinct().count()
    pending = total - delivered
    
    today = date.today()
    overdue = user_obligations.filter(
        due_date__lt=today,
        submissions__isnull=True
    ).count()
    
    # Performance por período
    this_month = user_obligations.filter(
        due_date__month=today.month,
        due_date__year=today.year
    )
    
    last_month = user_obligations.filter(
        due_date__month=today.month - 1 if today.month > 1 else 12,
        due_date__year=today.year if today.month > 1 else today.year - 1
    )
    
    # Performance por empresa
    company_performance = user_obligations.values('company__name').annotate(
        total=Count('id'),
        delivered=Count('id', filter=Q(submissions__isnull=False)),
        pending=Count('id', filter=Q(submissions__isnull=True)),
        overdue=Count('id', filter=Q(due_date__lt=today, submissions__isnull=True))
    ).order_by('-total')
    
    # Performance por tipo de obrigação
    obligation_performance = user_obligations.values('obligation_type__name', 'obligation_type__recurrence').annotate(
        total=Count('id'),
        delivered=Count('id', filter=Q(submissions__isnull=False)),
        pending=Count('id', filter=Q(submissions__isnull=True)),
        overdue=Count('id', filter=Q(due_date__lt=today, submissions__isnull=True))
    ).order_by('-total')
    
    # Tempo médio de entrega (dias após vencimento)
    delivered_obligations = user_obligations.filter(submissions__isnull=False)
    avg_delivery_delay = 0
    if delivered_obligations.exists():
        delays = []
        for obligation in delivered_obligations:
            submission = obligation.submissions.first()
            if submission:
                delay = (submission.delivery_date - obligation.due_date).days
                delays.append(delay)
        if delays:
            avg_delivery_delay = sum(delays) / len(delays)
    
    return Response({
        'user': {
            'id': user.id,
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email
        },
        'performance': {
            'total_obligations': total,
            'delivered': delivered,
            'pending': pending,
            'overdue': overdue,
            'compliance_rate': round((delivered / total * 100) if total > 0 else 0, 2),
            'avg_delivery_delay': round(avg_delivery_delay, 2)
        },
        'monthly_comparison': {
            'this_month': {
                'total': this_month.count(),
                'delivered': this_month.filter(submissions__isnull=False).distinct().count(),
                'pending': this_month.count() - this_month.filter(submissions__isnull=False).distinct().count()
            },
            'last_month': {
                'total': last_month.count(),
                'delivered': last_month.filter(submissions__isnull=False).distinct().count(),
                'pending': last_month.count() - last_month.filter(submissions__isnull=False).distinct().count()
            }
        },
        'by_company': list(company_performance),
        'by_obligation_type': list(obligation_performance)
    })
