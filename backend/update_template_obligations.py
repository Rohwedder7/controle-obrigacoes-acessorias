#!/usr/bin/env python
"""
Script para atualizar o template de obrigações com o campo CNPJ
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "Obrigações"

# Definir cabeçalhos
headers = [
    'CNPJ da Empresa',
    'Estado',
    'Tipo de Obrigação',
    'Nome da Obrigação',
    'Competência',
    'Data de Vencimento',
    'Prazo de Entrega',
    'Usuário Responsável',
    'Data Inicial de Validade',
    'Data Final de Validade',
    'Notas'
]

# Adicionar cabeçalhos
ws.append(headers)

# Formatar cabeçalho
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Ajustar largura das colunas
column_widths = [18, 8, 30, 30, 12, 15, 15, 20, 15, 15, 40]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + i)].width = width

# Congelar primeira linha
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{chr(65 + len(headers) - 1)}1"

# Salvar arquivo
output_path = os.path.join('media', 'templates', 'template_obrigacoes.xlsx')
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)

print(f"Template criado com sucesso em: {output_path}")
print(f"Cabeçalhos: {', '.join(headers)}")

